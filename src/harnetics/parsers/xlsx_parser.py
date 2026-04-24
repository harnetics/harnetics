# [INPUT]: 依赖 openpyxl 库、stdlib csv 模块与 models.document.Section
# [OUTPUT]: 对外提供 parse_xlsx(file_path, doc_id) 与 parse_csv(file_path, doc_id) -> list[Section]
# [POS]: parsers 包的 Excel/CSV 解析器，xlsx 按 Sheet 名称切分章节，csv 以文件名作单章节
# [PROTOCOL]: 变更时更新此头部，然后检查 AGENTS.md

from __future__ import annotations

import csv
from pathlib import Path

from harnetics.models.document import Section

# ================================================================
# Excel (.xlsx) 解析器
# ================================================================

def parse_xlsx(file_path: str, doc_id: str) -> list[Section]:
    """将 .xlsx 文件按 Sheet 名称切分为 Section 列表。

    策略：
    - 每个 Sheet 生成一个 Section
    - heading = Sheet 名称
    - content = 逐行拼接（非空单元格用制表符分隔，行间换行）
    - 完全空的行跳过
    """
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(str(Path(file_path)), read_only=True, data_only=True)
    sections: list[Section] = []

    for order_index, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append("\t".join(cells))
        content = "\n".join(rows)
        sections.append(
            Section(
                section_id=f"{doc_id}-sec-{order_index}",
                doc_id=doc_id,
                heading=sheet_name,
                content=content,
                level=1,
                order_index=order_index,
            )
        )

    wb.close()

    if not sections:
        return [
            Section(
                section_id=f"{doc_id}-sec-0",
                doc_id=doc_id,
                heading="(empty workbook)",
                content="",
                level=0,
                order_index=0,
            )
        ]

    return sections


# ================================================================
# CSV 解析器（DOORS 导出格式等）
# ================================================================

def parse_csv(file_path: str, doc_id: str) -> list[Section]:
    """将 .csv 文件解析为单个 Section。

    策略：
    - heading = 文件名（不含扩展名）
    - content = 每行转换为 "key: value" 格式字符串（DictReader）
    - 编码优先 UTF-8，失败降级 GBK
    """
    path = Path(file_path)
    text = _read_csv_text(path)

    import io
    reader = csv.DictReader(io.StringIO(text))
    lines: list[str] = []
    for row in reader:
        line_parts = [f"{k}: {v}" for k, v in row.items() if v is not None and str(v).strip()]
        if line_parts:
            lines.append(" | ".join(line_parts))

    content = "\n".join(lines)

    # ---- 若无列头，回退为逐行原始内容 ----
    if not lines:
        raw_reader = csv.reader(io.StringIO(text))
        content = "\n".join("\t".join(row) for row in raw_reader if any(c.strip() for c in row))

    return [
        Section(
            section_id=f"{doc_id}-sec-0",
            doc_id=doc_id,
            heading=path.stem,
            content=content,
            level=1,
            order_index=0,
        )
    ]


def _read_csv_text(path: Path) -> str:
    """尝试 UTF-8 读取，失败时降级 GBK。"""
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="gbk", errors="replace")
